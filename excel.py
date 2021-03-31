# import openpyxl as xl
# wb = xl.load_workbook("transactions.xlsx")
# sheet = wb["Sheet1"]
# cell = sheet['a1']
# cell = sheet.cell(1, 1)     # returns the same as above (row 1, column 1)
#
# for row in range(2,sheet.max_row + 1):
#     cell = sheet.cell(row, 3)           # the 3 stands for column 3
#     corrected_price = cell.value * 0.9
#     corrected_price_cell = sheet.cell(row, 4)       # the 4 stands for column 4
#     corrected_price_cell.value = corrected_price
#
#     wb.save('transactions2.xlsx')           # to save in another file trans2
#
#    # print(row)
#print(cell.value)
#print(sheet.max_row)        # would generate the number of rows

# --- TO ADD THE CHART ---------
import openpyxl as xl
from openpyxl.chart import BarChart, Reference      # the two are classes

def process_workbook(filename):
    wb = xl.load_workbook("filename")
    sheet = wb["Sheet1"]
    #cell = sheet['a1']
    #cell = sheet.cell(1, 1)     # returns the same as above (row 1, column 1)

    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row, 3)           # the 3 stands for column 3
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)       # the 4 stands for column 4
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,min_row=2, max_row=sheet.max_row, min_col=4,max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('filename')           # to save in another file transactions2
